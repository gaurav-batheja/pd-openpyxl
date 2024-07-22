import openpyxl
from fuzzywuzzy import process, fuzz
import pandas as pd

df =pd.read_excel("C:/Users/Gaurav/Desktop/spell.xlsx",sheet_name=["Sheet1","Sheet2"])
wb = openpyxl.load_workbook("C:/Users/Gaurav/Desktop/spell.xlsx")
sheet2_df=df["Sheet2"]
dict={}
ws1 = wb["Sheet1"]
ws2 = wb["Sheet2"]
row_no=1
for row_ws1 in (ws1.iter_rows()):
    cell_ws1=row_ws1[0].value
    max_ratio=0
    for row_ws2 in (ws2.iter_rows()):
        cell_ws2 = row_ws2[0].value
        curr_ratio=fuzz.ratio(cell_ws1,cell_ws2)
        if curr_ratio>max_ratio:
            max_ratio=curr_ratio
            match=(cell_ws2,cell_ws1)
    # =VLOOKUP(A2:A41,Sheet2!A2:B11,2,0)
    ws2.cell(column=1,row=row_no).value=cell_ws1
    row_no+=1
    dict[match[0]]=sheet2_df["id"]
    if row_no>ws2.max_row:
        break

print(ws2["A2"].value)
wb.save("C:/Users/Gaurav/Desktop/spell.xlsx")